#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
DailyBatchProcedure.py（改訂版）
目的:
  - JPX の component-file を自動検出してダウンロード（CSV/XLS/XLSX 自動判定）
  - プライム / グロース / スタンダード（内国株式）の行だけ処理
  - 各銘柄について irbank から基本指標とチャート（直近20営業日）を取得
  - Google News RSS でニュース感情スコアを取得
  - K列以降に指定の指標を追加して daily_analyze_YYYYMMDD.xls に保存（可能なら）
  - 出力後にニュース偏差値を計算して隣列に追加し、条件でフィルタした別ファイルを作成
方針:
  - 依存ライブラリは最小限（環境に無ければ CSV 出力にフォールバック）
  - 並列処理は使わない（逐次処理）
  - 個別 JSON は一時的に作成するが、処理終了後にすべて削除する
  - コメントはすべて日本語、平易で保守しやすい実装
"""

import os
import re
import csv
import json
import time
import math
import statistics
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime
import tempfile
import subprocess
import random

def get_volume_surge_codes():
    """
    松井証券の出来高急増ランキング(m-table)から銘柄コード一覧を取得する。
    _http_get_text_simple を使う実装。
    戻り値: ['1301', '7203', ...] のようなコード一覧
    """
    url = "https://finance.matsui.co.jp/ranking-volume-surge/index"
    # polite delay（固定秒は bot 判定されやすいのでランダム）
    time.sleep(1 + random.uniform(0, 1.5))
    html = _http_get_text_simple(url)
    # m-table を抽出
    table_match = re.search(
        r'<table[^>]*class="m-table"[^>]*>(.*?)</table>',
        html,
        re.DOTALL | re.IGNORECASE
    )

    if not table_match:
        raise RuntimeError("m-table が見つかりませんでした。HTML構造が変わった可能性があります。")
    table_html = table_match.group(1)
    # td の中から銘柄コード（4〜5桁の数字）を抽出
    codes = []
    for td in re.findall(r'<td[^>]*>(.*?)</td>', table_html, re.DOTALL | re.IGNORECASE):
        text = re.sub(r'<.*?>', '', td).strip()  # タグ除去
        # 改行で分割
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if len(lines) < 2:
            continue

        # 2行目に「150A 東G」などが入っている
        second_line = lines[1]

        # 先頭の英数字（銘柄コード）だけ抽出
        m = re.match(r'^([0-9A-Za-z]+)', second_line)
        if m:
            code = m.group(1)
            codes.append(code)
    return sorted(codes)
# ------------------------------------------------------------------
# 既存の irbank 関連関数を優先して読み込む
# 同じフォルダに irbank_basic.py / irbank_chart.py があればそれを使う
# 無ければ簡易フォールバック実装を使う（実運用では既存モジュールを用意してください）
# ------------------------------------------------------------------

# 簡易フォールバック（最低限の動作確認用）
def _http_get_text_simple(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    print(url + "へリクエストを送信します")
    with urllib.request.urlopen(req) as res:
        return res.read().decode("utf-8", errors="ignore")

def fetch_irbank_basic(code):
    """簡易フォールバック: 主要指標を可能な限り抽出（実運用では既存モジュールを使う）"""
    html = _http_get_text_simple(f"https://irbank.net/{code}")
    def find(p):
        m = re.search(p, html, re.DOTALL)
        return m.group(1) if m else None
    return {
        "code": code,
        "PER": find(r"PER（連）.*?class=\"text\">([\d\.,億万％\-%+]+)"),
        "PBR": find(r"PBR（連）.*?class=\"text\">([\d\.,億万％\-%+]+)"),
        "ROE": find(r"ROE（連）.*?class=\"text\">([\d\.,億万％\-%+]+)"),
        "ROA": find(r"ROA（連）.*?class=\"text\">([\d\.,億万％\-%+]+)"),
        "EPS": find(r"EPS（連）.*?class=\"text\">([\d\.,億万％\-%+]+)"),
        "BPS": find(r"BPS（連）.*?class=\"text\">([\d\.,億万％\-%+]+)"),
        # 出来高等はページ構造に依存するため見つからない場合あり
        "volume": find(r"出来高.*?class=\"co_sm\">([\d,億万]+)"),
        "price_5d": find(r"株価（5日）.*?class=\"text\">([\d,]+)"),
        "price_25d": find(r"株価（25日）.*?class=\"text\">([\d,]+)"),
        "volume_5d_value": find(r"出来高（5日）.*?class=\"co_sm\">([\d,]+)")
     }

def fetch_irbank_chart(code):
    """簡易フォールバック: tbc テーブルから直近20行を抜く"""
    html = _http_get_text_simple(f"https://irbank.net/{code}/chart")
    table_m = re.search(r'<table[^>]*id="tbc"[^>]*>.*?</table>', html, re.DOTALL)
    if not table_m:
        return []
    table_html = table_m.group(0)
    rows = re.findall(r'<tr.*?</tr>', table_html, re.DOTALL)
    out = []
    for row in rows[:20]:
        cells = re.findall(r'<td.*?>(.*?)</td>', row, re.DOTALL)
        clean = [re.sub(r'<.*?>', '', c).strip() for c in cells]
        if len(clean) < 10:
            continue
        out.append({
            "date": clean[0],
            "open": clean[1],
            "high": clean[2],
            "low": clean[3],
            "close": clean[4],
            "volume": clean[6],
            "market_cap": clean[7],
            "deviation_25d": clean[8],
            "PER": clean[9],
            "PBR": clean[10] if len(clean) > 10 else None
        })
        return out

# ------------------------------------------------------------------
# ニュース感情スコア（Google News RSS）
# ------------------------------------------------------------------
POSITIVE_WORDS = [
    "好調", "最高", "増益", "増配", "上昇", "高騰", "買い", "強気", "追い風", "成長",
    "拡大", "改善", "好転", "好材料", "回復", "プラス", "黒字", "好決算", "好評価",
    "ポジティブ", "高評価"
]
NEGATIVE_WORDS = [
    "不調", "悪化", "減益", "減配", "下落", "暴落", "売り", "弱気", "逆風", "縮小",
    "悪材料", "赤字", "マイナス", "減速", "低迷", "不振", "ネガティブ", "下方修正",
    "警戒", "懸念", "下降"
]
GOOGLE_NEWS_RSS = "https://news.google.com/rss/search?q={}&hl=ja&gl=JP&ceid=JP:ja"
QUERY_PREFIX = "株価 銘柄コード "

def sentiment_score(text: str) -> float:
    """テキストから感情スコアを 0-100 で返す（50 が中立）"""
    if not text or not text.strip():
        return 50.0
    normalized = text.lower()
    pos = sum(normalized.count(w) for w in POSITIVE_WORDS)
    neg = sum(normalized.count(w) for w in NEGATIVE_WORDS)
    total = pos + neg
    if total == 0:
        return 50.0
    raw = (pos - neg) / total
    score = ((raw + 1) / 2) * 100
    return round(score, 1)

def fetch_news_score(query: str, timeout=10):
    """Google News RSS を叩いて平均感情スコアを返す。失敗時は 50 を返す"""
    try:
        import requests
        url = GOOGLE_NEWS_RSS.format(requests.utils.quote(query))
        headers = {"User-Agent": "Python Sentiment Script"}
        print(url + "へリクエストを送信します")
        r = requests.get(url, headers=headers, timeout=timeout)
        r.raise_for_status()
        xml_text = r.text
    except Exception:
        # requests が無ければ urllib で取得
        try:
            q = urllib.request.quote(query, safe='')
            url = GOOGLE_NEWS_RSS.format(q)
            print(url + "へリクエストを送信します")
            req = urllib.request.Request(url, headers={"User-Agent": "Python Sentiment Script"})
            with urllib.request.urlopen(req, timeout=timeout) as res:
                xml_text = res.read().decode("utf-8", errors="ignore")
        except Exception:
            return 50.0

    try:
        root = ET.fromstring(xml_text)
        scores = []
        for item in root.findall(".//item"):
            title = item.findtext("title", "")
            desc = item.findtext("description", "")
            scores.append(sentiment_score(f"{title} {desc}"))
        if not scores:
            return 50.0
        return round(sum(scores) / len(scores), 1)
    except Exception:
        return 50.0

# ------------------------------------------------------------------
# 数値正規化ユーティリティ
#  - カンマ除去、%/倍の除去、億/万 の換算を行い数値（int/float）を返す
# ------------------------------------------------------------------
def normalize_number(value):
    """文字列や数値を受け取り、数値（int/float）または None を返す"""
    if value is None:
        return None
    v = str(value).strip()
    if v == "" or v in ["-", "―", "—", "N/A", "－"]:
        return None
    v = v.replace(",", "").replace(" ", "")
    is_percent = v.endswith("%")
    if is_percent:
        v = v[:-1]
    v = v.replace("倍", "")
    try:
        if "億" in v:
            v = v.replace("億", "")
            num = float(v) * 1e8
            return int(num) if num.is_integer() else float(num)
        if "万" in v:
            v = v.replace("万", "")
            num = float(v) * 1e4
            return int(num) if num.is_integer() else float(num)
        v_clean = re.sub(r"[^\d\.\-+eE]", "", v)
        if v_clean == "":
            return None
        num = float(v_clean)
        if math.isfinite(num) and num.is_integer():
            return int(num)
        return float(num)
    except Exception:
        return None

# ------------------------------------------------------------------
# JPX ファイルのダウンロード（component-file を自動検出）
# 戻り値: (local_path, ext) ext は 'csv'/'xls'/'xlsx'
# ------------------------------------------------------------------
def download_jpx_list(save_dir="."):
    page_url = "https://www.jpx.co.jp/markets/statistics-equities/misc/01.html"
    print(page_url + "へリクエストを送信します")
    req = urllib.request.Request(page_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as res:
        page_html = res.read().decode("utf-8", errors="ignore")

    m = re.search(r'<div[^>]*class="component-file"[^>]*>.*?href="([^"]+)"', page_html, re.DOTALL)
    if not m:
        raise RuntimeError("JPX のダウンロードリンクが見つかりません。")
    link = m.group(1)
    if link.startswith("/"):
        file_url = "https://www.jpx.co.jp" + link
    elif link.startswith("http"):
        file_url = link
    else:
        file_url = "https://www.jpx.co.jp/" + link.lstrip("./")

    req = urllib.request.Request(file_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as res:
        raw = res.read()

    lower = file_url.lower()
    if lower.endswith(".csv") or lower.endswith(".txt"):
        ext = "csv"
        local = os.path.join(save_dir, "jpx_list.csv")
    elif lower.endswith(".xls"):
        ext = "xls"
        local = os.path.join(save_dir, "jpx_list.xls")
    elif lower.endswith(".xlsx"):
        ext = "xlsx"
        local = os.path.join(save_dir, "jpx_list.xlsx")
    else:
        if raw[:2] == b'PK':
            ext = "xlsx"
            local = os.path.join(save_dir, "jpx_list.xlsx")
        else:
            ext = "csv"
            local = os.path.join(save_dir, "jpx_list.csv")

    with open(local, "wb") as f:
        f.write(raw)

    return local, ext

# ------------------------------------------------------------------
# LibreOffice (soffice) を使って xls/xlsx を CSV に変換する補助関数
# ------------------------------------------------------------------
def _convert_with_soffice(input_path, out_dir):
    cmd = ["soffice", "--headless", "--convert-to", "csv", "--outdir", out_dir, input_path]
    try:
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except FileNotFoundError:
        raise RuntimeError("LibreOffice (soffice) が見つかりません。soffice をインストールするか xlrd/openpyxl を導入してください。")
    except subprocess.CalledProcessError as e:
        raise RuntimeError(f"soffice による変換に失敗しました: {e}")

    base = os.path.basename(input_path)
    name, _ = os.path.splitext(base)
    candidate = os.path.join(out_dir, name + ".csv")
    if os.path.exists(candidate):
        return candidate

    files = sorted(
        (os.path.join(out_dir, p) for p in os.listdir(out_dir)),
        key=lambda p: os.path.getmtime(p),
        reverse=True
    )
    for f in files:
        if f.lower().endswith(".csv"):
            return f
    raise RuntimeError("soffice で変換した CSV が見つかりませんでした。")

# ------------------------------------------------------------------
# JPX ファイル解析（CSV は自動エンコーディング判定、XLS/XLSX は xlrd/openpyxl/soffice の順で対応）
# 戻り値: (header_list, rows_list)
# ------------------------------------------------------------------
def parse_jpx_file(path, ext):
    if ext == "csv":
        raw = open(path, "rb").read()
        text = None
        for enc in ("utf-8-sig", "utf-8", "shift_jis", "cp932", "euc_jp"):
            try:
                text = raw.decode(enc)
                break
            except Exception:
                text = None
        if text is None:
            raise RuntimeError("CSV のデコードに失敗しました（エンコーディング不明）。")
        rows = list(csv.reader(text.splitlines()))
        if not rows:
            return [], []
        return rows[0], rows[1:]

    if ext == "xls":
        try:
            import xlrd
            book = xlrd.open_workbook(path, formatting_info=False)
            sheet = book.sheet_by_index(0)
            header = sheet.row_values(0)
            rows = []
            for i in range(1, sheet.nrows):
                rows.append(sheet.row_values(i))
            return header, rows
        except ModuleNotFoundError:
            try:
                with tempfile.TemporaryDirectory() as td:
                    csv_path = _convert_with_soffice(path, td)
                    return parse_jpx_file(csv_path, "csv")
            except Exception as e:
                raise RuntimeError(
                    "XLS を読み取れません。対処方法:\n"
                    "  1) pip install xlrd==1.2.0\n"
                    "  2) LibreOffice をインストールして soffice を PATH に追加\n"
                    f"詳細: {e}"
                )
        except Exception as e:
            raise RuntimeError(f"XLS の読み込みでエラー: {e}")

    if ext == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            header = list(next(it))
            rows = [list(r) for r in it]
            return header, rows
        except ModuleNotFoundError:
            try:
                with tempfile.TemporaryDirectory() as td:
                    csv_path = _convert_with_soffice(path, td)
                    return parse_jpx_file(csv_path, "csv")
            except Exception as e:
                raise RuntimeError(
                    "XLSX を読み取れません。対処方法:\n"
                    "  1) pip install openpyxl\n"
                    "  2) LibreOffice をインストールして soffice を PATH に追加\n"
                    f"詳細: {e}"
                )
        except Exception as e:
            raise RuntimeError(f"XLSX の読み込みでエラー: {e}")

    raise RuntimeError("未対応のファイル形式です。")

# ------------------------------------------------------------------
# 対象銘柄フィルタ（B列=コード, D列=市場区分 を想定）
# 戻り値: list of (code_str, original_row)
# ------------------------------------------------------------------
def filter_target_rows(rows):
    targets = []
    for row in rows:
        if len(row) < 4:
            continue
        code_raw = row[1]
        market = row[3]  # D列（市場・商品区分）
        try:
            code = str(int(float(code_raw)))
        except Exception:
            code = str(code_raw).strip()
        if market in ["プライム（内国株式）", "グロース（内国株式）", "スタンダード（内国株式）"] and code in  get_volume_surge_codes():
            targets.append((code, row))
    return targets

# ------------------------------------------------------------------
# 1銘柄処理: irbank 取得、正規化、ニューススコア取得、行拡張
# 戻り値: (code, extended_row, norm_chart, error_str)
# 注意: 出力列は以下の順（列を詰めた仕様）
#   PER, PBR, ROE, ROA, EPS, BPS, chart_20days_json, news_positive_score, news_deviation_score, error
# news_deviation_score は後処理で追加するためここでは空で返す
# ------------------------------------------------------------------
def process_one(code, original_row):
    row = list(original_row)
    error_msg = ""
    basic = {}
    chart = []
    try:
        basic = fetch_irbank_basic(code) or {}
    except Exception as e:
        error_msg += f"basic_error:{e};"
    try:
        chart = fetch_irbank_chart(code) or []
    except Exception as e:
        error_msg += f"chart_error:{e};"

    try:
        news_score = fetch_news_score(QUERY_PREFIX + str(code))
    except Exception:
        news_score = 50.0

    # 基本指標のキーは環境によって異なる可能性があるため複数キーを許容して取得
    def get_basic_value(keys):
        for k in keys:
            if k in basic and basic.get(k) is not None:
                return basic.get(k)
        return None

    # 取得して正規化する（数値として出力）
    per = normalize_number(get_basic_value(["PER", "per"]))
    pbr = normalize_number(get_basic_value(["PBR", "pbr"]))
    roe = normalize_number(get_basic_value(["ROE", "roe"]))
    roa = normalize_number(get_basic_value(["ROA", "roa"]))
    eps = normalize_number(get_basic_value(["EPS", "eps"]))
    bps = normalize_number(get_basic_value(["BPS", "bps"]))

    # チャートを正規化（各要素）
    norm_chart = []
    for item in chart:
        nc = {
            "date": item.get("date"),
            "open": normalize_number(item.get("open")),
            "high": normalize_number(item.get("high")),
            "low": normalize_number(item.get("low")),
            "close": normalize_number(item.get("close")),
            "volume": normalize_number(item.get("volume")),
            "market_cap": normalize_number(item.get("market_cap")) if item.get("market_cap") else None,
            "deviation_25d": normalize_number(item.get("deviation_25d")),
            "PER": normalize_number(item.get("PER") or item.get("per")),
            "PBR": normalize_number(item.get("PBR") or item.get("pbr"))
        }
        norm_chart.append(nc)

    # チャート JSON とニューススコアを追加（偏差値は後で計算）
    chart_json = json.dumps(norm_chart, ensure_ascii=False)
    new_fields = [
        per, pbr, roe, roa, eps, bps,
        chart_json,
        news_score,
        None,        # news_deviation_score（後で埋める）
        error_msg
    ]

    row.extend(new_fields)
    return code, row, norm_chart, error_msg

# ------------------------------------------------------------------
# ヘッダ拡張（列を詰めた仕様）
# ------------------------------------------------------------------
def extend_header(header):
    new_columns = [
        "PER", "PBR", "ROE", "ROA", "EPS", "BPS",
        "chart_20days_json", "news_positive_score", "news_deviation_score", "error"
    ]
    return list(header) + new_columns

# ------------------------------------------------------------------
# 出力: まず .xls を試し、無ければ .xlsx、最終的に CSV にフォールバック
# - 依存ライブラリが無ければ CSV を出力（pip を最小限にする方針）
# ------------------------------------------------------------------
def write_output_file(header, rows, out_basename):
    # 1) try xlwt -> .xls
    try:
        import xlwt
        out_path = out_basename + ".xls"
        wb = xlwt.Workbook()
        ws = wb.add_sheet("sheet1")
        for c, h in enumerate(header):
            ws.write(0, c, h)
        for r, row in enumerate(rows, start=1):
            for c, v in enumerate(row):
                if isinstance(v, (dict, list)):
                    ws.write(r, c, json.dumps(v, ensure_ascii=False))
                else:
                    ws.write(r, c, "" if v is None else v)
        wb.save(out_path)
        return out_path
    except Exception:
        pass

    # 2) try openpyxl -> .xlsx
    try:
        import openpyxl
        from openpyxl import Workbook
        out_path = out_basename + ".xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        ws.append(header)
        for row in rows:
            ws.append([("" if v is None else v) for v in row])
        wb.save(out_path)
        return out_path
    except Exception:
        pass

    # 3) fallback CSV
    out_path = out_basename + ".csv"
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    return out_path

# ------------------------------------------------------------------
# 偏差値計算ユーティリティ（ニューススコアのリストから偏差値を返す）
# 偏差値 = 50 + 10 * ( (x - mean) / std )
# std が 0 の場合は 50 を返す
# ------------------------------------------------------------------
def compute_deviation_scores(values):
    """values: list of numbers -> returns list of 偏差値 (float)"""
    if not values:
        return []
    # None を除外して平均と標準偏差を計算するが、元の順序を保つ
    nums = [v for v in values if v is not None]
    if not nums:
        return [50.0] * len(values)
    mean = statistics.mean(nums)
    std = statistics.pstdev(nums) if len(nums) > 0 else 0.0
    out = []
    for v in values:
        if v is None:
            out.append(50.0)
        else:
            if std == 0:
                out.append(50.0)
            else:
                z = (v - mean) / std
                dev = 50.0 + 10.0 * z
                out.append(round(dev, 1))
    return out

# ------------------------------------------------------------------
# フィルタ条件: 直近営業日の出来高が過去5日平均の2.5倍以上かつ前営業日の2倍以上
# - norm_chart は日付降順（最新が最初）を想定するが、日付解析できる場合はソートする
# - 戻り値: True/False
# ------------------------------------------------------------------
def volume_spike_condition(norm_chart):
    """
    norm_chart: list of dict with 'date' and 'volume' (数値)
    条件:
      v0 = 最新日の出来高
      v_prev = 1日前の出来高
      avg5 = 過去5日平均（v_prev を含めない、v1..v5 とする）
    判定: v0 >= 2.5 * avg5 and v0 >= 2 * v_prev
    """
    if not norm_chart or len(norm_chart) < 6:
        return False
    # 日付が文字列なら試しにソート。日付形式が不明なら配列順を尊重
    def parse_date(d):
        if not d:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(d, fmt)
            except Exception:
                continue
        return None

    # try to sort by date descending if dates parseable
    parsed = []
    for item in norm_chart:
        dt = parse_date(item.get("date"))
        parsed.append((dt, item))
    if all(p[0] is not None for p in parsed):
        parsed.sort(key=lambda x: x[0], reverse=True)
        sorted_items = [p[1] for p in parsed]
    else:
        # fallback: assume input order is newest first
        sorted_items = norm_chart

    # extract volumes as numbers, skip None
    vols = [item.get("volume") for item in sorted_items]
    # need at least 6 numeric volumes
    numeric_vols = []
    for v in vols:
        if isinstance(v, (int, float)):
            numeric_vols.append(v)
        else:
            try:
                numeric_vols.append(float(v))
            except Exception:
                numeric_vols.append(None)
    # ensure we have at least 6 valid numbers at top
    if len(numeric_vols) < 6:
        return False
    # take first 6 entries
    top6 = numeric_vols[:6]
    if any(v is None for v in top6):
        return False
    v0 = top6[0]
    v_prev = top6[1]
    # avg of next 5 days after v0: top6[1]..top6[5] -> but spec says "過去5日平均" likely excluding v0, so use top6[1:6]
    avg5 = sum(top6[1:6]) / 5.0
    if avg5 == 0:
        return False
    if v0 >= 2.5 * avg5 and v0 >= 2.0 * v_prev:
        return True
    return False

# ------------------------------------------------------------------
# メイン処理（逐次処理）
# - 個別 JSON は一時的に作成するが、処理終了後にすべて削除する
# - 出力後にニュース偏差値を計算して隣列に追加し、条件でフィルタした別ファイルを作成
# ------------------------------------------------------------------
def run_daily_batch(output_dir="."):
    today = datetime.now().strftime("%Y%m%d")
    out_basename = os.path.join(output_dir, f"daily_analyze_{today}")

    print("JPX の銘柄一覧をダウンロードします...")
    local_path, ext = download_jpx_list()
    print(f"ダウンロード完了: {local_path} (ext={ext})")

    print("ファイルを解析します...")
    header, rows = parse_jpx_file(local_path, ext)
    print(f"総行数: {len(rows)}")

    targets = filter_target_rows(rows)
    print(f"対象銘柄数 (プライム/グロース/スタンダード): {len(targets)}")

    out_header = extend_header(header)
    out_rows = []

    # 一時 JSON ファイルを記録しておくリスト（後で削除）
    temp_json_files = []

    # 逐次処理（保守性重視）
    for code, row in targets:
        try:
            _, result_row, norm_chart, err = process_one(code, row)
            out_rows.append(result_row)
            # 個別 JSON を一時保存（後で削除）
            json_path = f"irbank_{code}.json"
            with open(json_path, "w", encoding="utf-8") as jf:
                json.dump({"code": code, "chart_20days": norm_chart}, jf, ensure_ascii=False, indent=2)
            temp_json_files.append(json_path)
            print(f"[OK] {code}")
            # サイト負荷を下げるため短いスリープ
            time.sleep(10 + random.uniform(0, 20))
        except Exception as e:
            print(f"[ERR] {code} -> {e}")

    # まずは全データを出力（偏差値は未計算のまま）
    print("全データを出力します（偏差値は後で追加）...")
    out_path = write_output_file(out_header, out_rows, out_basename)
    print("一次出力完了:", out_path)

    # 偏差値計算: news_positive_score 列の値を抽出して偏差値を計算し、各行の隣列に埋める
    # out_rows の各行は元の行 + [PER,PBR,ROE,ROA,EPS,BPS,chart_json,news_score,news_dev(None),error]
    news_scores = []
    for r in out_rows:
        # news score はヘッダ末尾から3番目の位置（index = original_len + 7）
        # safer: find index by header name
        # compute index:
        try:
            idx = out_header.index("news_positive_score")
        except ValueError:
            idx = len(out_header) - 3
        val = r[idx] if idx < len(r) else None
        try:
            news_scores.append(float(val) if val is not None else None)
        except Exception:
            news_scores.append(None)

    news_dev_list = compute_deviation_scores(news_scores)

    # ニュース偏差値を各行に書き込む
    try:
        dev_idx = out_header.index("news_deviation_score")
    except ValueError:
        dev_idx = len(out_header) - 2
    for i, r in enumerate(out_rows):
        # ensure row is long enough
        while len(r) < len(out_header):
            r.append(None)
        r[dev_idx] = news_dev_list[i]

    # 上書きで偏差値入りのファイルを出力（上書き）
    print("偏差値を埋めたファイルを出力します...")
    out_path_with_dev = write_output_file(out_header, out_rows, out_basename)
    print("偏差値埋め込み出力完了:", out_path_with_dev)

    # フィルタ処理:
    # ① 直近営業日の出来高が過去5日平均の2.5倍以上かつ前営業日の2倍以上
    # ② ニュース偏差値 >= 56
    filtered_rows = []
    # find indices for needed columns
    try:
        chart_idx = out_header.index("chart_20days_json")
    except ValueError:
        chart_idx = None
    try:
        dev_idx = out_header.index("news_deviation_score")
    except ValueError:
        dev_idx = None

    for r in out_rows:
        keep = True
        # ニュース偏差値条件
        if dev_idx is None:
            keep = False
        else:
            dev_val = r[dev_idx]
            try:
                if dev_val is None or float(dev_val) < 56.0:
                    keep = False
            except Exception:
                keep = False
        # 出来高スパイク条件
        if keep:
            if chart_idx is None:
                keep = False
            else:
                chart_json = r[chart_idx]
                try:
                    norm_chart = json.loads(chart_json) if chart_json else []
                except Exception:
                    norm_chart = []
                # norm_chart should have 'volume' numeric; ensure conversion
                for item in norm_chart:
                    if "volume" in item:
                        item["volume"] = normalize_number(item.get("volume"))
                if not volume_spike_condition(norm_chart):
                    keep = False
        if keep:
            filtered_rows.append(r)

    # 出力ファイル名（フィルタ後）
    out_basename_filtered = os.path.join(output_dir, f"daily_analyze_{today}_filtered")
    print(f"フィルタ後の行数: {len(filtered_rows)}")
    out_path_filtered = write_output_file(out_header, filtered_rows, out_basename_filtered)
    print("フィルタ後出力完了:", out_path_filtered)

    # 一時 JSON をすべて削除する（存在確認してから削除）
    for p in temp_json_files:
        try:
            if os.path.exists(p):
                os.remove(p)
        except Exception:
            print(f"一時ファイルの削除に失敗しました: {p}")

    print("一時ファイルを削除しました。処理終了。")
    return out_path_with_dev, out_path_filtered

# ------------------------------------------------------------------
# 実行
# ------------------------------------------------------------------
if __name__ == "__main__":
    #get_volume_surge_codes()
    run_daily_batch()
